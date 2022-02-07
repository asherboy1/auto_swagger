

from openpyxl import load_workbook


class ExcelHandler:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
       # self.wb = load_workbook()

    def open(self):
        self.wb = load_workbook(self.file_name)
        if isinstance(self.sheet_name, int):
            self.sheet = self.wb.worksheets[self.sheet_name]
        else:
            self.sheet = self.wb[self.sheet_name]

    # def choose_sheet(self, sheet_name):
    #     """选择表单.
    #     sheet_name 是整数，根据索引获取。
    #     如果是字符串，根据名字获取 '20190920'
    #     """
    #     if isinstance(sheet_name, int):
    #         return self.wb.worksheets[sheet_name]
    #     return self.wb[sheet_name]

    def headers(self):
        """获取标题"""
        self.open()
        headers = [c.value for c in self.sheet[1]]
        self.wb.close()
        return headers

    def read(self,start_row=2, start_column=1):
        """获取所有的数据"""
        self.open()
        sheet = self.sheet

        header = [c.value for c in sheet[1]]

        data = []
        for row in range(start_row, sheet.max_row + 1):
            row_data = []
            for column in range(start_column, sheet.max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            row_data = dict(zip(header, row_data))
            data.append(row_data)
        self.wb.close()
        return data

    def save(self):
        """保存"""
        self.wb.save(self.file_name)
        self.wb.close()

    def write(self, row, column, data):
        self.open()
        self.sheet.cell(row, column).value = data
        self.save()



if __name__ == '__main__':
    xls = ExcelHandler('cases1.xlsx',"code")
    # print(xls.read( 2, 1 ))